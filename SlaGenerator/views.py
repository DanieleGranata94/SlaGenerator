import collections
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

import neo4j
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from neo4j import GraphDatabase
from project.settings import neo4jUser, neo4jPass, neo4jURI

neo4jUsername = neo4jUser
neo4jPassword = neo4jPass
neo4jUri = neo4jURI

from SlaGenerator.models import MACM, Asset, Relation, Protocol, Attribute, Attribute_value, Asset_Attribute_value, \
    Threat_Attribute_value, Threat_CIA, Threat_Stride, ThreatAgentQuestion, Reply, TAReplies_Question, TAReplyCategory, \
    ThreatAgentCategory, TACategoryAttribute, ThreatAgentAttribute, ThreatAgentRiskScores, StrideImpactRecord, Stride, \
    MACM_ThreatAgent


def threat_agent_wizard(request,appId):
    context={}
    #Generate question and related replies
    questions=ThreatAgentQuestion.objects.all()
    questions_replies=TAReplies_Question.objects.all()
    questions_replies_list=[]
    for question in questions:
        replies = []
        question_replies_dict = {}
        for reply in questions_replies:
            if question==reply.question:
                replies.append(reply.reply.reply)
        question_replies_dict['question']=question.question
        question_replies_dict['replies']=replies
        questions_replies_list.append(question_replies_dict)
    context['questions_replies']=questions_replies_list
    context['appId']=appId
    return render(request, 'threat_agent_wizard.html', context)

@csrf_exempt
def threat_agent_generation(request, appId):
    print(appId)
    context={}
    ThreatAgents = []
    ThreatAgentsPerAsset = []
    #for category in ThreatAgentCategory.objects.all():   #inizializzo la lista finale a tutti i TA
        #ThreatAgents.append(category)

    for reply in request.POST: #per ogni risposta al questionario
        ReplyObject=Reply.objects.filter(reply=reply).get()
        tareplycategories=TAReplyCategory.objects.filter(reply=ReplyObject)

        TAList=[]
        for replycategory in tareplycategories.all(): #ogni categoria relativa ad una singola risposta
            #print(replycategory.reply.reply + " "+ replycategory.category.category)
            TAList.append(replycategory.category)
            question = TAReplies_Question.objects.filter(reply=ReplyObject)
        ThreatAgentsPerAsset.append((TAList,question))

    numQ3=0
    numQ4=0
    #conto il numero di risposte date per Q3 e Q4
    for ThreatAgentsList,question in ThreatAgentsPerAsset: #per ogni risposta
        questionId=question.get().question.Qid
        if(questionId=="Q3"):
            numQ3+=1
        if(questionId=="Q4"):
            numQ4+=1

    i=0
    j=0
    ThreatAgentsListTemp=[]
    for ThreatAgentsList,question in ThreatAgentsPerAsset: #per ogni risposta
        questionId=question.get().question.Qid
        if(questionId=="Q1"):
            ThreatAgents=ThreatAgentsList
        if(questionId=="Q2"):
            ThreatAgents=intersection(ThreatAgents,ThreatAgentsList)
        if(questionId=="Q3"):
            if(i==0):
                ThreatAgentsListTemp = ThreatAgentsList
            elif(i<numQ3):
                ThreatAgentsList=union(ThreatAgentsList,ThreatAgentsListTemp)
                ThreatAgentsListTemp=ThreatAgentsList
            if(i==numQ3-1):
                ThreatAgents = intersection(ThreatAgents, ThreatAgentsList)
            i = i + 1

        if(questionId=="Q4"):
            if(j==0):
                ThreatAgentsListTemp=ThreatAgentsList
                j=j+1
            elif(j==1):
                ThreatAgentsListTemp = ThreatAgentsList
                j=j+1
            elif(j<numQ4):
                ThreatAgentsList=union(ThreatAgentsList,ThreatAgentsListTemp)
                ThreatAgentsListTemp=ThreatAgentsList

    ThreatAgents = intersection(ThreatAgents, ThreatAgentsList)
    ThreatAgentsWithInfo={}
    for ta in ThreatAgents:
        ThreatAgentsWithInfo[ta]=list(TACategoryAttribute.objects.filter(category=ta))
        MACM_ThreatAgent.objects.get_or_create(
            app = MACM.objects.get(appId=appId),
            category=ta
        )



    context={'ThreatAgents':ThreatAgentsWithInfo}
    context['appId']=appId
    return render(request, 'threat_agent_generation.html',context=context)

def intersection(lst1, lst2):
    lst3 = [value for value in lst1 if value in lst2]
    return lst3

def union(lst1, lst2):
    lst3 = list(set(lst1+lst2))
    return lst3

@csrf_exempt
def calculate_threat_agent_risks(request,appId):
    OWASP_Motive_TOT = 0
    OWASP_Size_TOT = 0
    OWASP_Opportunity_TOT = 0
    OWASP_Skill_TOT = 0
    sommapesi = 0

    for category,risk_value in request.POST.items():
        TACategory=ThreatAgentCategory.objects.get(category=category)
        #per ogni categoria ottieni i Attribute relativi e calcola i 4 parametri owasp con le formule nella tesi.
        TACategoryAttributes=TACategoryAttribute.objects.filter(category=TACategory)
        OWASP_Motive=0
        OWASP_Size=0
        OWASP_Opportunity=0
        OWASP_Skill=0
        limits=0
        intent=0
        access=0
        resources=0
        visibility=0
        skills=0

        OWASP_Motives=[]

        #scorro gli attributi di category
        for TACategoryAttributeVar in TACategoryAttributes:
            if(TACategoryAttributeVar.attribute.attribute=='Skills'):
                skills=TACategoryAttributeVar.attribute.score
            if(TACategoryAttributeVar.attribute.attribute=='Resources'):
                resources=TACategoryAttributeVar.attribute.score
            if (TACategoryAttributeVar.attribute.attribute == 'Visibility'):
                visibility= TACategoryAttributeVar.attribute.score
            if (TACategoryAttributeVar.attribute.attribute == 'Limits'):
                limits= TACategoryAttributeVar.attribute.score
            if (TACategoryAttributeVar.attribute.attribute == 'Intent'):
                intent= TACategoryAttributeVar.attribute.score
            if (TACategoryAttributeVar.attribute.attribute == 'Access'):
                access= TACategoryAttributeVar.attribute.score

        if(risk_value=='L'):
            risk_valueNum= 1
        if (risk_value == 'M'):
            risk_valueNum = 2
        if (risk_value == 'H'):
            risk_valueNum = 3



        sommapesi=sommapesi+risk_valueNum
        OWASP_Motive= ((((intent/2)+(limits/4))/2) * 10)
        OWASP_Opportunity= ((((access/2)+(resources/6)+(visibility/4))/3) * 10)
        OWASP_Size= (resources/6) * 10
        OWASP_Skill= (skills/4) * 10

        OWASP_Motive_TOT += (OWASP_Motive * risk_valueNum)
        OWASP_Opportunity_TOT += OWASP_Opportunity * risk_valueNum
        OWASP_Size_TOT += OWASP_Size * risk_valueNum
        OWASP_Skill_TOT += OWASP_Skill * risk_valueNum

    OWASP_Skill_TOT= int(round(OWASP_Skill_TOT/sommapesi))
    OWASP_Motive_TOT= int(round(OWASP_Motive_TOT/sommapesi))
    OWASP_Size_TOT= int(round(OWASP_Size_TOT/sommapesi))
    OWASP_Opportunity_TOT= int(round(OWASP_Opportunity_TOT/sommapesi))

    app=MACM.objects.get(appId=appId)

    ScoreAlreadyCreated=ThreatAgentRiskScores.objects.filter(app=app)
    if(not ThreatAgentRiskScores.objects.filter(app=app).exists()):
        obj=ThreatAgentRiskScores.objects.get_or_create(
        app=app,
        skill=OWASP_Skill_TOT,
        size = OWASP_Size_TOT,
        motive = OWASP_Motive_TOT,
        opportunity = OWASP_Opportunity_TOT)

    return render(request, 'stride_impact_evaluation.html', {"appId": appId})


@csrf_exempt
def stride_impact_evaluation(request,appId):
    threats_list = threat_modeling(appId)
    stride_impact_list=[]
    app = MACM.objects.get(appId=appId)
    if (not StrideImpactRecord.objects.filter(app=app).exists()):
        save=False
        count=0
        for info,value in (request.POST).items():
            splittedInfo=info.split('_')
            impactValues=[]
            stride=splittedInfo[0]
            impactInfo=splittedInfo[1]
            #print(stride+" "+impactInfo)
            if(stride=='spoofing'):
                strideCategory='Spoofing'
            if(stride=='tampering'):
                strideCategory='Tampering'
            if(stride=='reputation'):
                strideCategory='Reputation'
            if(stride=='informationdisclosure'):
                strideCategory='Information Disclosure'
            if(stride=='dos'):
                strideCategory='Denial Of Services'
            if(stride=='elevationofprivileges'):
                strideCategory='Elevation of privileges'

            if(impactInfo=='noncompliance'):
                NonComplianceString='Non Compliance'
                NonComplianceValue=value
                stride_impact_list.append((strideCategory,NonComplianceString,NonComplianceValue))
            if(impactInfo=='financialdamage'):
                FinancialDamageValue=value
                FinancialDamageString='Financial Damage'
                stride_impact_list.append((strideCategory,FinancialDamageString,FinancialDamageValue))
            if(impactInfo=='reputationdamage'):
                ReputationDamageValue=value
                ReputationDamageString='Reputation Damage'
                stride_impact_list.append((strideCategory,ReputationDamageString,ReputationDamageValue))
            if(impactInfo=='privacyviolation'):
                PrivacyViolationValue=value
                PrivacyViolationString='Privacy Violation'
                stride_impact_list.append((strideCategory,PrivacyViolationString,PrivacyViolationValue))

            count+=1
            if(count==4):
                save=True
            if(save):
                strideObject = Stride.objects.get(category=strideCategory)
                strideImpactRecord = StrideImpactRecord.objects.all().get_or_create(app=app,
                                                        stride=strideObject,
                                                        financialdamage=FinancialDamageValue,
                                                        reputationdamage=ReputationDamageValue,
                                                        noncompliance=NonComplianceValue,
                                                        privacyviolation=PrivacyViolationValue)
                save=False
                count=0

    return render(request, 'threat_modeling.html', {"threats": threats_list, "appId": appId,'stride_impact_list':stride_impact_list})


def apps_management(request):
    ordered_apps = []
    context = {}
    try:
        graphDriver = GraphDatabase.driver(uri=neo4jUri, auth=(neo4jUsername, neo4jPassword))
        session = graphDriver.session()
        nodes_string = session.run("match (node) return node")
        nodes = [record for record in nodes_string.data()]
        apps = {}
        for node in nodes:
            try:
                apps[node['node']['app_id']] = node['node']['application']
            except:
                print("Cannot parse graph with node " + str(node['node']))
        ordered_apps = collections.OrderedDict(sorted(apps.items()))
        # print(ordered_apps)
        for appId, application in ordered_apps.items():
            MACM_instance = MACM(appId=appId, application=application)
            MACMvalue = MACM.objects.all().filter(appId=appId, application=application)
            if not MACMvalue:
                MACM_instance.save()
            graphDriver.close()
        context = {
            'apps': ordered_apps
        }
    except neo4j.exceptions.ServiceUnavailable as error:
        print(error)
        context = {
            'error': error
        }
    return render(request, 'apps_management.html', context)


def get_graphNodesbyAppId(appId):
    graph = GraphDatabase.driver(uri=neo4jUri, auth=(neo4jUsername, neo4jPassword))
    session = graph.session()
    nodes_string = session.run("MATCH (node { app_id:  \'" + str(appId) + "\' }) RETURN "
                                                                          "node,labels(node) as nodeType")
    nodes = [record for record in nodes_string.data()]
    session.close()
    return nodes


def get_graphRelationbyAppId(appId):
    graph = GraphDatabase.driver(uri=neo4jUri, auth=(neo4jUsername, neo4jPassword))
    session = graph.session()
    nodes_string = session.run("MATCH (client { app_id:  \'" + str(
        appId) + "\' }) -[relation]->(server) RETURN client,labels(client) as clientType,"
                 " relation,TYPE(relation) as relationType,relation.protocol as protocol, "
                 "server,labels(server) as serverType")
    nodes = [record for record in nodes_string.data()]
    session.close()
    return nodes


def macm_viewer(request, appId):

    return render(request, 'macm_viewer.html')

@csrf_exempt
def risk_analysis(request, appId):
    app = MACM.objects.get(appId=appId)
    appName=app.application
    SelectedComponentName=''
    componentsWithThreats=[]
    components=Asset.objects.filter(app=app)

    try:
        if request.POST['dropdown']:
            SelectedComponentName=request.POST['dropdown']
        else:
            SelectedComponentName=components[0].name
    except:
        print()
    componentUnderAnalysis=components[0]
    for component in components:
        if (len(threat_modeling_per_assetFun(component.id)) != 0):
            if(SelectedComponentName==component.name):
                componentsWithThreats.append((component,True))
                componentUnderAnalysis=component
            else:
                componentsWithThreats.append((component,False))

    threats = threat_modeling_per_assetFun(componentUnderAnalysis.id)

    TAscores=ThreatAgentRiskScores.objects.filter(app=app)

    #ricerca ultimo risultato.
    maxtimeTA = TAscores[0].updated_at
    lastScore=TAscores[0]
    for Tascore in TAscores:
        if(Tascore.updated_at>maxtimeTA):
            lastScore=Tascore

    SIRecords = StrideImpactRecord.objects.filter(app=app)

    PreCondition="[n,n,n]"
    PostCondition="[n,n,n]"
    LossOfConfidentiality=0
    LossOfIntegrity=0
    LossOfAvailability=0
    LossOfCPostConditionValue = 0
    LossOfIPostConditionValue = 0
    LossOfAPostConditionValue = 0
    LossOfCPreConditionValue = 0
    LossOfIPreConditionValue = 0
    LossOfAPreConditionValue = 0

    for threat in threats:
        print(threat[0].source)
        PreCondition=str(threat[0].PreCondition)
        PostCondition=str(threat[0].PostCondition)
        maxFinancial = 0
        maxReputation = 0
        maxnoncompliance = 0
        maxprivacy = 0
        for SIRecord in SIRecords:
            for Threatstride in threat[1]:
                if(SIRecord.stride.category.lower()==Threatstride.lower()):
                    if(maxFinancial < SIRecord.financialdamage):
                        maxFinancial=SIRecord.financialdamage
                    if (maxReputation < SIRecord.reputationdamage):
                        maxReputation = SIRecord.reputationdamage
                    if (maxnoncompliance < SIRecord.noncompliance):
                        maxnoncompliance = SIRecord.noncompliance
                    if (maxprivacy < SIRecord.privacyviolation):
                        maxprivacy = SIRecord.privacyviolation
        threat[0].financial=maxFinancial
        threat[0].reputation=maxReputation
        threat[0].noncompliance=maxnoncompliance
        threat[0].privacy=maxprivacy

        #elimino [ e ]
        print(PreCondition)
        print(PostCondition)
        try:

            PreCondition.replace("[","")
            PreCondition.replace("]","")
            PostCondition.replace("[","")
            PostCondition.replace("]","")



            #splitto con le ,
            PreCondition=PreCondition.split(",")
            PostCondition=PostCondition.split(",")


            if(PreCondition[0]=='n'):
                LossOfCPreConditionValue=0
            if (PreCondition[0] == 'p'):
                LossOfCPreConditionValue = 1
            if(PreCondition[0]=='f'):
                LossOfCPreConditionValue=2

            if(PostCondition[0]=='n'):
                LossOfCPostConditionValue=0
            if (PostCondition[0] == 'p'):
                LossOfCPostConditionValue = 1
            if(PostCondition[0]=='f'):
                LossOfCPostConditionValue=2

            LossOfConfidentiality=((LossOfCPostConditionValue+LossOfCPreConditionValue)*3)+1

            if (PreCondition[1] == 'n'):
                LossOfIPreConditionValue = 0
            if (PreCondition[1] == 'p'):
                LossOfIPreConditionValue = 1
            if (PreCondition[1] == 'f'):
                LossOfIPreConditionValue = 2

            if (PostCondition[1] == 'n'):
                LossOfIPostConditionValue = 0
            if (PostCondition[1] == 'p'):
                LossOfIPostConditionValue = 1
            if (PostCondition[1] == 'f'):
                LossOfIPostConditionValue = 2

            LossOfIntegrity = ((LossOfIPostConditionValue + LossOfIPreConditionValue) * 3) + 1

            if (PreCondition[2] == 'n'):
                LossOfAPreConditionValue = 0
            if (PreCondition[2] == 'p'):
                LossOfAPreConditionValue = 1
            if (PreCondition[2] == 'f'):
                LossOfAPreConditionValue = 2

            if (PostCondition[2] == 'n'):
                LossOfAPostConditionValue = 0
            if (PostCondition[2] == 'p'):
                LossOfAPostConditionValue = 1
            if (PostCondition[2] == 'f'):
                LossOfAPostConditionValue = 2

            LossOfAvailability = ((LossOfAPostConditionValue + LossOfAPreConditionValue) * 3) + 1

            threat[0].lossofc=LossOfConfidentiality
            threat[0].lossofi=LossOfIntegrity
            threat[0].lossofa=LossOfAvailability

        except:
            print("iNFO MISSING")








    return render(request, 'risk_analysis.html', {"appName": appName,"ComponentName":SelectedComponentName,"threats":threats,
                                                  "components":componentsWithThreats,"ThreatAgentScores":lastScore})





def asset_management(request, appId):
    # save assets info in sqlite
    # nodes = Asset.objects.all().filter(app=MACM.objects.get(appId=appId))
    # metto nodes=None perchè così prende sempre fa neo4j (dovrei gestire la coerenza fra i due DB)
    nodes = None
    # connect to neo4j only if sqlite assets are empty (API are laggy)
    if not nodes:
        nodes = get_graphNodesbyAppId(appId)
        for node in nodes:
            # print(node["node"]["name"]+" "+ node["node"]["type"])
            # print(node)

            asset = Asset.objects.all().get_or_create(app=MACM.objects.get(appId=appId),
                                                      name=node["node"]["name"])
            # mi salvo id sqlite in dizionario
            node['id'] = asset[0].id

            try:
                # vedo se il nome del componente è un attribute value
                # per il 5g andrebbero considerate le minacce sia di SERVICE.Web che di UE (ad esempio)
                Attribute_value_instance = Attribute_value.objects.get(attribute_value=node["node"]["type"])
                Asset_Attribute_value.objects.all().get_or_create(asset=asset[0],
                                                                  attribute_value=Attribute_value_instance)
                nodes = Asset_Attribute_value.objects.all().filter(app=MACM.objects.get(appId=appId))
            except:
                print()

    # save relation info in sqlite

    arches = get_graphRelationbyAppId(appId)
    for arch in arches:
        Asset_client = Asset.objects.all().filter(name=arch["client"]["name"], app=MACM.objects.get(appId=appId))
        Asset_server = Asset.objects.all().filter(name=arch["server"]["name"], app=MACM.objects.get(appId=appId))

        if arch["protocol"] is None:
            print()
        elif isinstance(arch["protocol"], str):
            # single protocol in one arch
            try:
                Relation.objects.all().get_or_create(asset=Asset.objects.get(name=arch["client"]["name"],
                                                                             app=MACM.objects.get(appId=appId)),
                                                     protocol=Protocol.objects.get(protocol=arch["protocol"]),
                                                     app=MACM.objects.get(appId=appId),
                                                     relation_type=arch["relationType"],
                                                     role="client")
                Relation.objects.all().get_or_create(asset=Asset.objects.get(name=arch["server"]["name"],
                                                                             app=MACM.objects.get(appId=appId)),
                                                     protocol=Protocol.objects.get(protocol=arch["protocol"]),
                                                     app=MACM.objects.get(appId=appId),
                                                     relation_type=arch["relationType"],
                                                     role="server")
            except:
                print()
                # print("Protocol info not found in arch between " + str(arch["client"]["name"]) + " and " + str(
                # arch["server"]["name"]))
        elif isinstance(arch["protocol"], list):
            for protocol in arch["protocol"]:
                # print(protocol)
                # multiple protocol in one arch
                try:
                    Relation.objects.all().get_or_create(asset=Asset.objects.get(name=arch["client"]["name"],
                                                                                 app=MACM.objects.get(appId=appId)),
                                                         protocol=Protocol.objects.get(protocol=protocol),
                                                         app=MACM.objects.get(appId=appId),
                                                         relation_type=arch["relationType"],
                                                         role="client")
                    Relation.objects.all().get_or_create(asset=Asset.objects.get(name=arch["server"]["name"],
                                                                                 app=MACM.objects.get(appId=appId)),
                                                         protocol=Protocol.objects.get(protocol=protocol),
                                                         app=MACM.objects.get(appId=appId),
                                                         relation_type=arch["relationType"],
                                                         role="server")
                except:
                    print()
                    # print("Protocol info not found in arch between " + str(arch["client"]["name"]) + " and " + str(
                    #   arch["server"]["name"]))
            else:
                print("error getting protocol information")
        # we consider only relations with some associated properties
        relations = Relation.objects.all().filter(app=MACM.objects.get(appId=appId))
    return render(request, 'asset_management.html', {
        'nodes': nodes,
        'relations': relations,
        'appId': appId
    })

def threat_modeling_per_asset(request, appId, assetId):
    threats = []
    try:
        asset = Asset.objects.all().filter(id=assetId)[0]
        asset_attribute_value = Asset_Attribute_value.objects.all().filter(asset_id=assetId)
        threats_attribute_values = Threat_Attribute_value.objects.all().filter(
            attribute_value_id=asset_attribute_value[0].attribute_value.id)
        for threat_attribute_value in threats_attribute_values:
            strides_per_threat = []
            affectedRequirements = []
            try:
                for stride in Threat_Stride.objects.all().filter(threat=threat_attribute_value.threat):
                    strides_per_threat.append(stride.stride.category)
                for requirement in Threat_CIA.objects.all().filter(threat=threat_attribute_value.threat):
                    affectedRequirements.append(requirement.cia.requirement)
            except:
                print("Error in selecting additional info")

            threats.append((threat_attribute_value.threat, strides_per_threat, affectedRequirements))
    except:
        print("OutOfRange")

    return render(request, 'threat_modeling_per_asset.html', {
        'threats': threats,
        'asset': asset}
                  )

def threat_modeling_per_assetFun(assetId):
    threats = []
    try:
        asset = Asset.objects.all().filter(id=assetId)[0]
        asset_attribute_value = Asset_Attribute_value.objects.all().filter(asset_id=assetId)
        threats_attribute_values = Threat_Attribute_value.objects.all().filter(
            attribute_value_id=asset_attribute_value[0].attribute_value.id)
        for threat_attribute_value in threats_attribute_values:
            strides_per_threat = []
            affectedRequirements = []
            try:
                for stride in Threat_Stride.objects.all().filter(threat=threat_attribute_value.threat):
                    strides_per_threat.append(stride.stride.category)
                for requirement in Threat_CIA.objects.all().filter(threat=threat_attribute_value.threat):
                    affectedRequirements.append(requirement.cia.requirement)
            except:
                print("Error in selecting additional info")

            threats.append((threat_attribute_value.threat, strides_per_threat, affectedRequirements))
    except:
        print("OutOfRange")
    return threats

def threat_modeling(appId):
    threats_list = []
    nodes = get_graphNodesbyAppId(appId)
    for node in nodes:
        asset = Asset.objects.all().filter(name=node["node"]["name"])[0]
        asset_attribute_value = Asset_Attribute_value.objects.all().filter(asset=asset)
        try:
            #print(asset.name + " " + asset_attribute_value[0].attribute_value.attribute_value)
            threats_attribute_values = Threat_Attribute_value.objects.all().filter(
                attribute_value=asset_attribute_value[0].attribute_value)
            for threat_attribute_value in threats_attribute_values:
                strides_per_threat = []
                affectedRequirements = []
                try:
                   #print(Threat_Stride.objects.all().filter(threat=threat_attribute_value.threat))
                    for stride in Threat_Stride.objects.all().filter(threat=threat_attribute_value.threat):
                        strides_per_threat.append(stride.stride.category)
                    for requirement in Threat_CIA.objects.all().filter(threat=threat_attribute_value.threat):
                        affectedRequirements.append(requirement.cia.requirement)
                    threats_list.append((threat_attribute_value.threat, strides_per_threat, affectedRequirements, asset.name,
                     threat_attribute_value.attribute_value))
                except:
                    print("Error in selecting additional info")
        except:
            print()
    return threats_list

def export_threat_modeling(request, appId):
    if request.method == "POST":

        # help: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}-TM-report.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
            name=MACM.objects.get(appId=appId).application.replace(" ", "_")
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Threat_modeling_REPORT'
        columns = ['#','Threat Agent','Asset name', 'Asset type', 'Threat',  'CIA', 'STRIDE','Behaviour']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(name="Times New Roman", size=12, bold=True, color='FF0000')
            cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                 right=Side(border_style="thin", color='FF000000'),
                                 top=Side(border_style="thin", color='FF000000'),
                                 bottom=Side(border_style="thin", color='FF000000'), )

        threats_list = threat_modeling(appId)
        ThreatAgents=(MACM_ThreatAgent.objects.all().filter(app=MACM.objects.get(appId=appId)))
        for ta in ThreatAgents:
            for threat in threats_list:
                row_num += 1
                stride=""
                cia=""
                for index,single in enumerate(threat[1]):
                    if not index==len(threat[1])-1:
                        stride+=single+", "
                    else:
                        stride+=single

                for index,single in enumerate(threat[2]):
                    if not index==len(threat[1])-1:
                        cia+=single+", "
                    else:
                        cia+=single

                # columns = ['Asset name', 'Asset type', 'Threat', 'Description', 'CIA', 'STRIDE']
                print(threat[4].attribute_value)
                # Define the data for each cell in the row
                row = [
                    row_num,
                    ta.category.category,
                    threat[3],
                    threat[4].attribute_value,
                    threat[0].name,
                    cia,
                    stride,
                    threat[0].description,
                ]

                # Assign the data for each cell of the row
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                    cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                         right=Side(border_style="thin", color='FF000000'),
                                         top=Side(border_style="thin", color='FF000000'),
                                         bottom=Side(border_style="thin", color='FF000000'), )

                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
                        cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                        cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                             right=Side(border_style="thin", color='FF000000'),
                                             top=Side(border_style="thin", color='FF000000'),
                                             bottom=Side(border_style="thin", color='FF000000'), )
            # Per effettuare il resize delle celle in base a quella più grande
            dims = {}

            from openpyxl.styles import Alignment

            for row in worksheet.rows:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))+0.05
            for col, value in dims.items():
                worksheet.column_dimensions[col].width = value


        workbook.save(response)

        return response
